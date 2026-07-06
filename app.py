import streamlit as st
import os
import json
import base64
import requests
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Gemini OCR 測試", layout="centered")

# ── 側邊欄：Gemini 模型選擇 ──────────────────────────────────────
with st.sidebar:
    st.header("⚙️ 設定")
    gemini_model = st.selectbox(
        "Gemini 模型",
        options=[
            "gemini-3.5-flash",
            "gemini-3.1-flash-lite",
        ],
        index=1,
        format_func=lambda x: {
            "gemini-3.5-flash": "Gemini 3.5 Flash",
            "gemini-3.1-flash-lite": "Gemini 3.1 Flash Lite",
        }[x],
    )
    st.caption(f"目前選用：`{gemini_model}`")

st.markdown("<h2 style='font-size: 20px; margin-bottom: 8px;'>輥輪組裝報表 OCR 分析</h2>", unsafe_allow_html=True)

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")

def create_excel_report(parsed_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "軋輥組裝報表"

    font_header = Font(name="微軟正黑體", size=10, bold=True, color="FFFFFF")
    font_body = Font(name="微軟正黑體", size=10)
    font_bold = Font(name="微軟正黑體", size=10, bold=True)
    
    fill_header_main = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    fill_pair_1 = PatternFill(start_color="16A085", end_color="16A085", fill_type="solid")
    fill_pair_2 = PatternFill(start_color="1ABC9C", end_color="1ABC9C", fill_type="solid")
    fill_row_even = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid") 
    
    align_center = Alignment(horizontal="center", vertical="center")
    border_thin = Border(
        left=Side(style='thin', color='D5D8DC'), right=Side(style='thin', color='D5D8DC'),
        top=Side(style='thin', color='D5D8DC'), bottom=Side(style='thin', color='D5D8DC')
    )

    # 1. 寫入第一列標題 (擴充至 30 欄並設定合併)
    for col in range(1, 31):
        cell = ws.cell(row=1, column=col)
        cell.font = font_header
        cell.alignment = align_center
        cell.border = border_thin

    ws.cell(row=1, column=1, value="施工項目").fill = fill_header_main
    ws.cell(row=1, column=2, value="型號").fill = fill_header_main
    
    for i in range(7):
        col_id = 3 + i * 4
        col_val = 5 + i * 4
        current_pair_fill = fill_pair_1 if i % 2 == 0 else fill_pair_2
        
        ws.cell(row=1, column=col_id, value="編號")
        ws.merge_cells(start_row=1, start_column=col_id, end_row=1, end_column=col_id+1)
        
        ws.cell(row=1, column=col_val, value="尺寸")
        ws.merge_cells(start_row=1, start_column=col_val, end_row=1, end_column=col_val+1)
        
        for c in range(col_id, col_id + 2):
            ws.cell(row=1, column=c).fill = current_pair_fill
        for c in range(col_val, col_val + 2):
            ws.cell(row=1, column=c).fill = current_pair_fill

    # 2. 資料寫入與施工項目替換
    categories_order = [
        ("本體銲補", "再生"),
        ("軸頸銲補", "軸位再生"),
        ("本體未再生車修", "粗車"),
        ("軸頸未再生車修", "軸位粗車"),
        ("本體再生車修", "精車"),
        ("軸頸再生車修", "軸位精車"),
        ("真圓度", "圓度")
    ]

    current_row = 2

    for display_cat, json_cat in categories_order:
        if json_cat not in parsed_data or not parsed_data[json_cat]:
            continue
            
        model_dict = parsed_data[json_cat]
        is_first_cat_row = True
        
        for model, rollers in model_dict.items():
            if not rollers:
                continue
                
            items = list(rollers.items())
            
            for chunk_idx in range(0, len(items), 7):
                chunk = items[chunk_idx:chunk_idx+7]
                
                # 初始化 30 欄框線 (取消底色)
                for col in range(1, 31):
                    c = ws.cell(row=current_row, column=col)
                    c.border = border_thin
                
                cell_a = ws.cell(row=current_row, column=1)
                if is_first_cat_row:
                    cell_a.value = display_cat
                    cell_a.font = font_bold
                    is_first_cat_row = False
                cell_a.alignment = align_center
                
                cell_b = ws.cell(row=current_row, column=2)
                if chunk_idx == 0:
                    cell_b.value = model
                    cell_b.font = font_bold
                cell_b.alignment = align_center
                
                # 填入有資料的編號與尺寸，並合併
                for pair_idx, (r_id, r_val) in enumerate(chunk):
                    col_id = 3 + pair_idx * 4
                    col_val = 5 + pair_idx * 4
                    
                    cell_id = ws.cell(row=current_row, column=col_id, value=r_id)
                    ws.merge_cells(start_row=current_row, start_column=col_id, end_row=current_row, end_column=col_id+1)
                    
                    has_decimal = '.' in str(r_val)
                    try:
                        val_num = float(r_val) if has_decimal else int(r_val)
                    except ValueError:
                        val_num = r_val
                        
                    cell_val = ws.cell(row=current_row, column=col_val, value=val_num)
                    ws.merge_cells(start_row=current_row, start_column=col_val, end_row=current_row, end_column=col_val+1)
                    
                    cell_id.font = font_body
                    cell_id.alignment = align_center
                    cell_val.font = font_body
                    cell_val.alignment = align_center
                    
                    if isinstance(val_num, (int, float)) and has_decimal:
                        cell_val.number_format = '0.00'
                    elif isinstance(val_num, int):
                        cell_val.number_format = '0'
                
                # 💡 這裡就是新增的邏輯：把剩下的空白組也合併起來
                for empty_idx in range(len(chunk), 7):
                    empty_col_id = 3 + empty_idx * 4
                    empty_col_val = 5 + empty_idx * 4
                    
                    # 合併編號的空白格
                    ws.merge_cells(start_row=current_row, start_column=empty_col_id, end_row=current_row, end_column=empty_col_id+1)
                    # 合併尺寸的空白格
                    ws.merge_cells(start_row=current_row, start_column=empty_col_val, end_row=current_row, end_column=empty_col_val+1)

                current_row += 1

    # 3. 自動調整欄寬
    for col in ws.columns:
        max_width = 0
        for cell in col:
            if cell.value:
                text = str(cell.value)
                cell_width = sum(2.2 if ord(c) > 255 else 1.1 for c in text)
                if cell_width > max_width:
                    max_width = cell_width
        
        col_idx = col[0].column
        padding = 0.5 if col_idx == 1 else 3
        
        if col_idx >= 3:
            final_width = (max_width + padding) / 2
        else:
            final_width = max_width + padding
            
        ws.column_dimensions[get_column_letter(col_idx)].width = max(final_width, 5)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

uploaded_file = st.file_uploader(
    "上傳報表圖檔", type=["jpg", "jpeg", "png", "pdf"]
)

if uploaded_file:
    file_content = uploaded_file.getvalue()

    st.subheader("🟢 Gemini 結構化辨識")
    if not GEMINI_API_KEY:
        st.error("⚠️ 請設定 GEMINI_API_KEY 環境變數。")
    else:
        PROMPT = """
你是工廠軋輥維修報表的資料擷取助手。
請分析這張「軋輥組裝報表」圖片，依照以下規則輸出 JSON：

規則：
1. 掃描每一筆軋輥記錄（每一列）。注意報表中有區段標示不同的「輥輪型號」（例如 30D, 30S, 30L）。
2. 欄位對應如下：
   - 粗車（Roll 粗車尺寸，非軸位）
   - 再生（Roll 再生尺寸）
   - 精車（Roll 精車尺寸）
   - 軸位粗車（軸位粗車尺寸）
   - 軸位再生（軸位再生尺寸）
   - 軸位精車（軸位精車尺寸）
   - 圓度（對應表上最右側的「圓度」數值）

3. 若該格為數字（含小數），代表有施做，請記錄該數字（字串格式）。
4. 若該格為「X」或空白，代表未施做，請略過（不要包含在輸出中）。
5. 每個類別只輸出「有施做」的項目。
6. JSON 結構必須為三層：【施工類別】 -> 【輥輪型號】 -> 【輥輪編號: 尺寸】。

輸出格式範例（只輸出 JSON，不要加任何說明文字）：
{
  "粗車": {
    "30D": { "N30DL90": "288", "30DL13": "288" }
  },
  "再生": {
    "30D": { "N30DL90": "307" }
  },
  "精車": {
    "30D": { "N30DL90": "300.06" }
  },
  "軸位粗車": {
    "30S": { "V30S58": "146" }
  },
  "軸位再生": {
    "30S": { "V30S58": "155" }
  },
  "軸位精車": {
    "30S": { "V30S58": "149.97" }
  },
  "圓度": {
    "30D": { "N30DL90": "0.06", "30DL13": "0.06" }
  }
}

請直接輸出 JSON，不要加 markdown 代碼區塊。
""".strip()

        with st.spinner(f"使用 {gemini_model} 辨識中…"):
            try:
                fname = uploaded_file.name.lower()
                if fname.endswith(".pdf"):
                    mime = "application/pdf"
                elif fname.endswith(".png"):
                    mime = "image/png"
                else:
                    mime = "image/jpeg"

                b64_data = base64.b64encode(file_content).decode("utf-8")

                payload = {
                    "contents": [
                        {
                            "parts": [
                                {
                                    "inline_data": {
                                        "mime_type": mime,
                                        "data": b64_data,
                                    }
                                },
                                {"text": PROMPT},
                            ]
                        }
                    ]
                }

                api_url = (
                    f"https://generativelanguage.googleapis.com/v1beta/models/"
                    f"{gemini_model}:generateContent?key={GEMINI_API_KEY}"
                )
                resp = requests.post(
                    api_url,
                    headers={"Content-Type": "application/json"},
                    json=payload,
                    timeout=60,
                )
                resp.raise_for_status()
                raw_text = (
                    resp.json()
                    .get("candidates", [{}])[0]
                    .get("content", {})
                    .get("parts", [{}])[0]
                    .get("text", "")
                    .strip()
                )

                if raw_text.startswith("```"):
                    raw_text = raw_text.split("```")[1]
                    if raw_text.startswith("json"):
                        raw_text = raw_text[4:]
                    raw_text = raw_text.strip()

                parsed = json.loads(raw_text)
                st.success("✅ Gemini 辨識成功！")

                # ── 新增：匯出 Excel 下載按鈕 ──────────────────────
                excel_data = create_excel_report(parsed)
                st.download_button(
                    label="📥 下載 Excel 報表",
                    data=excel_data,
                    file_name="軋輥組裝報表整理.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )
                # ──────────────────────────────────────────────────
                CATEGORIES = ["粗車", "再生", "精車", "軸位粗車", "軸位再生", "軸位精車", "圓度"]
                for cat in CATEGORIES:
                    types_dict = parsed.get(cat, {})
                    total_count = sum(len(rollers) for rollers in types_dict.values())
                    
                    label = f"**{cat}**（{total_count} 件）" if total_count else f"{cat}（無施做）"
                    with st.expander(label, expanded=(total_count > 0)):
                        if total_count > 0:
                            for r_type, rollers in types_dict.items():
                                if rollers:
                                    st.markdown(f"**🔹 型號：{r_type}**")
                                    for roller_id, size in rollers.items():
                                        st.write(f"　- `{roller_id}` → **{size}**")
                        else:
                            st.caption("本次無此項目。")

                st.divider()
                st.caption("📋 原始 JSON")
                st.json(parsed)

            except json.JSONDecodeError as e:
                st.error(f"JSON 解析失敗：{e}")
                st.code(raw_text, language="json")
            except Exception as e:
                st.error(f"Gemini 辨識錯誤：{e}")
