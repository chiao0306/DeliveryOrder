import os
import json
import base64
import requests
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.responses import RedirectResponse
from pydantic import BaseModel
from typing import Dict, Any

# ==========================================
# 模組 1：影像辨識模組
# ==========================================
class ReportOCRRecognizer:
    """處理圖檔與 Gemini API 溝通的 OCR 辨識模組"""

    PROMPT = """
你是工廠軋輥維修報表的資料擷取助手。
請分析這張「軋輥組裝報表」圖片，依照以下規則輸出 JSON：

規則：
1. 掃描每一筆軋輥記錄（每一列）。注意報表中有區段標示不同的「輥輪型號」（例如 30D, 30S, 30L, 200, 170）。
2. 欄位對應如下：
   - 粗車（Roll 粗車尺寸，非軸位）
   - 再生（Roll 再生尺寸）
   - 精車（Roll 精車尺寸）
   - 軸位粗車（軸位粗車尺寸）
   - 軸位再生（軸位再生尺寸）
   - 軸位精車（軸位精車尺寸）
   - 圓度（對應表上最右側的「圓度」數值）
   - 軸位數量（對應表上的「軸位數量」欄位，若出現數字「2」請務必記錄）
3. 若該格為數字（含小數），代表有施做，請記錄該數字（字串格式）。
4. 若該格為「X」或空白，代表未施做，請略過（不要包含在輸出中）。
5. 每個類別只輸出「有施做」的項目。
6. JSON 結構必須為三層：【施工類別】 -> 【輥輪型號】 -> 【輥輪編號: 尺寸或數量】。
7. 【重要】若遇到「重複的輥輪編號」（也就是同一個編號在表中出現兩次以上），請務必在 JSON 的編號後方加上底線與流水號（例如遇到兩筆 ABC01，請輸出 "ABC01_1", "ABC01_2"），確保鍵值（Key）唯一，否則資料會被覆蓋遺失。此規則適用所有欄位。

請直接直接輸出 JSON，不要加 markdown 代碼區塊。
    """.strip()

    def __init__(self, api_key: str, model: str):
        if not api_key:
            raise ValueError("未設定 GEMINI_API_KEY 環境變數。")
        self.api_key = api_key
        self.model = model
        self.api_url = f"https://generativelanguage.googleapis.com/v1beta/models/{self.model}:generateContent?key={self.api_key}"

    def _get_mime_type(self, filename: str) -> str:
        fname = filename.lower()
        if fname.endswith(".pdf"):
            return "application/pdf"
        elif fname.endswith(".png"):
            return "image/png"
        return "image/jpeg"

    def _encode_image(self, file_content: bytes) -> str:
        return base64.b64encode(file_content).decode("utf-8")

    def _build_payload(self, mime_type: str, b64_data: str) -> dict:
        return {
            "contents": [
                {
                    "parts": [
                        {"inline_data": {"mime_type": mime_type, "data": b64_data}},
                        {"text": self.PROMPT},
                    ]
                }
            ],
            "generationConfig": {
                "thinkingConfig": {"thinkingLevel": "minimal"}
            },
        }

    def analyze(self, file_content: bytes, filename: str) -> str:
        mime_type = self._get_mime_type(filename)
        b64_data = self._encode_image(file_content)
        payload = self._build_payload(mime_type, b64_data)

        resp = requests.post(
            self.api_url,
            headers={"Content-Type": "application/json"},
            json=payload,
            timeout=120,
        )
        resp.raise_for_status()

        raw_text = resp.json().get("candidates", [{}])[0].get("content", {}).get("parts", [{}])[0].get("text", "").strip()
        return raw_text

# ==========================================
# 模組 2：資料清洗與轉換模組
# ==========================================
class ReportDataProcessor:
    """負責將 OCR 字串清洗並轉置為標準 JSON 結構"""

    def __init__(self):
        self.standard_cats = ["粗車", "再生", "精車", "軸位粗車", "軸位再生", "軸位精車", "圓度", "軸位數量"]

    def process(self, raw_text: str) -> dict:
        """主入口：執行清洗、解析與轉置"""
        clean_text = self._strip_markdown(raw_text)
        parsed_data = json.loads(clean_text)
        return self._normalize(parsed_data)

    def _strip_markdown(self, text: str) -> str:
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
        return text.strip()

    def _normalize(self, parsed_data: dict) -> dict:
        if any(cat in parsed_data for cat in self.standard_cats):
            return parsed_data
            
        raw_data = parsed_data
        if "軋輥維修報表" in parsed_data:
            raw_data = parsed_data["軋輥維修報表"]
        elif len(parsed_data) == 1:
            first_key = list(parsed_data.keys())[0]
            raw_data = parsed_data[first_key]

        normalized = {cat: {} for cat in self.standard_cats}
        
        if isinstance(raw_data, dict):
            for model, rollers in raw_data.items():
                if not isinstance(rollers, dict): 
                    continue
                for r_id, operations in rollers.items():
                    if not isinstance(operations, dict): 
                        continue
                    for op_name, value in operations.items():
                        if op_name in normalized:
                            if model not in normalized[op_name]:
                                normalized[op_name][model] = {}
                            normalized[op_name][model][r_id] = str(value)
                        
        return normalized

# ==========================================
# 模組 3：Excel 報表產出模組
# ==========================================
class ExcelReportGenerator:
    """負責將標準化 JSON 資料渲染並匯出為 Excel"""

    def __init__(self, parsed_data: dict):
        self.parsed_data = parsed_data
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "軋輥組裝報表"
        
        self.font_header = Font(name="微軟正黑體", size=10, bold=True, color="FFFFFF")
        self.font_body = Font(name="微軟正黑體", size=10)
        self.font_bold = Font(name="微軟正黑體", size=10, bold=True)
        self.fill_header_main = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        self.fill_pair_1 = PatternFill(start_color="16A085", end_color="16A085", fill_type="solid")
        self.fill_pair_2 = PatternFill(start_color="1ABC9C", end_color="1ABC9C", fill_type="solid")
        self.fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.align_center = Alignment(horizontal="center", vertical="center")

    def _get_border(self, col: int, is_last_row: bool = False) -> Border:
        """動態產生框線，自動消除合併儲存格的內部線殘影，並加上特定粗線"""
        left = Side(style='thin', color='D5D8DC')
        right = Side(style='thin', color='D5D8DC')
        top = Side(style='thin', color='D5D8DC')
        bottom = Side(style='medium', color='000000') if is_last_row else Side(style='thin', color='D5D8DC')
        
        # 處理合併儲存格：第3欄以後，每兩欄為一個合併單元(3-4為編號, 5-6為尺寸)
        # 將內部的框線設為 None，避免在底部粗線上產生「凸出來一根根」的殘影
        if col >= 3:
            if col % 2 == 1:  # 單數欄 (3, 5, 7...) 是合併儲存格的左半邊，清空右框線
                right = Side(border_style=None)
            else:             # 雙數欄 (4, 6, 8...) 是合併儲存格的右半邊，清空左框線
                left = Side(border_style=None)
                
        # 針對 型號(2) 及 每個完整施工項目的結尾(6, 10, 14, 18, 22, 26, 30) 加上右側粗直線
        if col in [2, 6, 10, 14, 18, 22, 26, 30]:
            right = Side(style='medium', color='000000')
            
        return Border(left=left, right=right, top=top, bottom=bottom)

    def _setup_headers(self):
        for col in range(1, 31):
            cell = self.ws.cell(row=1, column=col)
            cell.font = self.font_header
            cell.alignment = self.align_center
            cell.border = self._get_border(col, is_last_row=False)

        self.ws.cell(row=1, column=1, value="施工項目").fill = self.fill_header_main
        self.ws.cell(row=1, column=2, value="型號").fill = self.fill_header_main
        
        for i in range(7):
            col_id = 3 + i * 4
            col_val = 5 + i * 4
            current_pair_fill = self.fill_pair_1 if i % 2 == 0 else self.fill_pair_2
            
            self.ws.cell(row=1, column=col_id, value="編號")
            self.ws.merge_cells(start_row=1, start_column=col_id, end_row=1, end_column=col_id+1)
            self.ws.cell(row=1, column=col_val, value="尺寸")
            self.ws.merge_cells(start_row=1, start_column=col_val, end_row=1, end_column=col_val+1)
            
            for c in range(col_id, col_id + 2):
                self.ws.cell(row=1, column=c).fill = current_pair_fill
            for c in range(col_val, col_val + 2):
                self.ws.cell(row=1, column=c).fill = current_pair_fill

    def _write_data(self):
        categories_order = [
            ("本體銲補", "再生"),
            ("軸頸銲補", "軸位再生"),
            ("本體未再生車修", "粗車"),
            ("軸頸未再生車修", "軸位粗車"),
            ("本體再生車修", "精車"),
            ("軸頸再生車修", "軸位精車"),
            ("真圓度", "圓度")
        ]

        current_row = 2

        for display_cat, json_cat in categories_order:
            if json_cat not in self.parsed_data or not self.parsed_data[json_cat]:
                continue
                
            model_dict = self.parsed_data[json_cat]
            is_first_cat_row = True
            cat_start_row = current_row 
            
            for model, rollers in model_dict.items():
                if not rollers:
                    continue
                    
                items = []
                is_shaft_cat = json_cat in ["軸位再生", "軸位粗車", "軸位精車"]
                qty_data = self.parsed_data.get("軸位數量", {}).get(model, {})
                
                for r_id, r_val in rollers.items():
                    display_id = r_id.split('_')[0] 
                    if is_shaft_cat and str(qty_data.get(r_id)) == "2":
                        items.append((display_id, r_val, True))
                        items.append((display_id, r_val, True))
                    else:
                        items.append((display_id, r_val, False))
                        
                if not items:
                    continue
                
                for chunk_idx in range(0, len(items), 7):
                    chunk = items[chunk_idx:chunk_idx+7]
                    
                    for col in range(1, 31):
                        c = self.ws.cell(row=current_row, column=col)
                        c.border = self._get_border(col, is_last_row=False)
                    
                    cell_a = self.ws.cell(row=current_row, column=1)
                    if is_first_cat_row:
                        cell_a.value = display_cat
                        cell_a.font = self.font_bold
                        is_first_cat_row = False
                    cell_a.alignment = self.align_center
                    
                    cell_b = self.ws.cell(row=current_row, column=2)
                    if chunk_idx == 0:
                        cell_b.value = model
                        cell_b.font = self.font_bold
                    cell_b.alignment = self.align_center
                    
                    for pair_idx, (r_id, r_val, is_yellow) in enumerate(chunk):
                        col_id = 3 + pair_idx * 4
                        col_val = 5 + pair_idx * 4
                        
                        cell_id = self.ws.cell(row=current_row, column=col_id, value=r_id)
                        self.ws.merge_cells(start_row=current_row, start_column=col_id, end_row=current_row, end_column=col_id+1)
                        
                        has_decimal = '.' in str(r_val)
                        try:
                            val_num = float(r_val) if has_decimal else int(r_val)
                        except ValueError:
                            val_num = r_val
                            
                        cell_val = self.ws.cell(row=current_row, column=col_val, value=val_num)
                        self.ws.merge_cells(start_row=current_row, start_column=col_val, end_row=current_row, end_column=col_val+1)
                        
                        cell_id.font = self.font_body
                        cell_id.alignment = self.align_center
                        cell_val.font = self.font_body
                        cell_val.alignment = self.align_center
                        
                        if isinstance(val_num, (int, float)) and has_decimal:
                            cell_val.number_format = '0.00'
                        elif isinstance(val_num, int):
                            cell_val.number_format = '0'
                            
                        if is_yellow:
                            for c in range(col_id, col_id + 2):
                                self.ws.cell(row=current_row, column=c).fill = self.fill_yellow
                            for c in range(col_val, col_val + 2):
                                self.ws.cell(row=current_row, column=c).fill = self.fill_yellow
                    
                    for empty_idx in range(len(chunk), 7):
                        empty_col_id = 3 + empty_idx * 4
                        empty_col_val = 5 + empty_idx * 4
                        self.ws.merge_cells(start_row=current_row, start_column=empty_col_id, end_row=current_row, end_column=empty_col_id+1)
                        self.ws.merge_cells(start_row=current_row, start_column=empty_col_val, end_row=current_row, end_column=empty_col_val+1)

                    current_row += 1
            
            # 在大項目結束時，將最後一列的底部設為粗線
            if current_row > cat_start_row:
                last_data_row = current_row - 1
                for col in range(1, 31):
                    cell = self.ws.cell(row=last_data_row, column=col)
                    cell.border = self._get_border(col, is_last_row=True)

    def _auto_fit_columns(self):
        for col in self.ws.columns:
            max_width = 0
            for cell in col:
                if cell.value:
                    text = str(cell.value)
                    cell_width = sum(2.2 if ord(c) > 255 else 1.1 for c in text)
                    if cell_width > max_width:
                        max_width = cell_width
            
            col_idx = col[0].column
            padding = 0.5 if col_idx == 1 else 3
            
            if col_idx >= 3:
                final_width = (max_width + padding) / 2
            else:
                final_width = max_width + padding
                
            self.ws.column_dimensions[get_column_letter(col_idx)].width = max(final_width, 5)

    def export(self) -> io.BytesIO:
        self._setup_headers()
        self._write_data()
        self._auto_fit_columns()
        
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output

# ==========================================
# FastAPI 伺服器與路由
# ==========================================
app = FastAPI(title="Gemini OCR API")

app.mount("/static", StaticFiles(directory="static"), name="static")

@app.get("/")
async def root():
    return RedirectResponse(url="/static/index.html")

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")

@app.post("/api/ocr")
async def analyze_ocr(file: UploadFile = File(...), model: str = Form(...)):
    if not GEMINI_API_KEY:
        raise HTTPException(status_code=500, detail="未設定 GEMINI_API_KEY 環境變數。")

    content = await file.read()
    
    try:
        recognizer = ReportOCRRecognizer(api_key=GEMINI_API_KEY, model=model)
        raw_text = recognizer.analyze(content, file.filename)
        
        processor = ReportDataProcessor()
        normalized_data = processor.process(raw_text)

        return JSONResponse(content=normalized_data)

    except json.JSONDecodeError:
        raise HTTPException(status_code=500, detail="JSON 解析失敗。")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Gemini 辨識錯誤: {str(e)}")

@app.post("/api/export-excel")
async def export_excel(data: Dict[str, Any]):
    try:
        generator = ExcelReportGenerator(data)
        excel_file = generator.export()
        
        headers = {
            'Content-Disposition': 'attachment; filename="report.xlsx"',
            'Access-Control-Expose-Headers': 'Content-Disposition'
        }
        return StreamingResponse(
            excel_file, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
            headers=headers
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8080))
    uvicorn.run(app, host="0.0.0.0", port=port)
